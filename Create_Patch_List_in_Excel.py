import datetime as dt
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import DEFAULT_FONT, Font
import sys
import re
import requests
from bs4 import BeautifulSoup
import PMS_Data as pmsd
import time
from tqdm import tqdm

undecidedList = []
removeList = []

startPeriod = None
endPeriod = None

global lastID

previousIDDic = {}
passiveIDDic = {}
passiveUpdateVersionDic = {}
previousPatchList = []
importantSet = set()
criticalSet = set()
exclusionPatchList = []
duplicationPatchList = []
newPatchList = []
totalFileVersionHistoryDic = {}


def getPatchPeriod():
    global startPeriod
    global endPeriod
    today = dt.date.today()
    beforeOneMonth = today - relativedelta(months=1)
    startPeriod = getPatchDateByMonth(beforeOneMonth)
    endPeriod = getPatchDateByMonth(today)


def getPatchDateByMonth(dateTime):
    patchDate = dateTime.replace(day=1)
    weeks = 0
        
    while weeks < 2:
        if patchDate.weekday() == 1:
            weeks += 1
        patchDate += dt.timedelta(days=1)

    return patchDate


def isPatchExclusion(des):
    return any(one in des for one in pmsd.patchExclusionList)

def isExceptionToInclude(des):
    for one in pmsd.exceptionToIncludeList:
        regexPattern = re.compile(one)
        result = re.findall(regexPattern, des)
        length = len(result)
        if length == 0:
            continue
        elif length == 1:
            return True
        else:
            return False
    return False

def isNumber(str):
    try:
        float(str)
        return True
    except ValueError:
        return False


def validatePatchInfo(guid, kbid, des):
    if (not isNumber(kbid)) or kbid == 0:
        exclusionPatchList.append([guid, kbid, des])
        return False
    if (not isinstance(des, str)) or isPatchExclusion(des):
        exclusionPatchList.append([guid, kbid, des])
        return False
    if (guid+'\t'+str(int(kbid))+'\n') in previousPatchList:
        duplicationPatchList.append([guid, kbid, des])
        return False
    return True


def makePreviousIDDic():
    global previousIDDic
    global startPeriod
    
    f = open('./' + startPeriod.strftime('%Y_%m_%d') + '_Previous_Passive_Update.txt', 'r')
    line = f.readline()
    tempList = line.split('=')
    previousIDDic = eval(tempList[1])
    

def makePassiveUpdateVersionDic():
    global passiveUpdateVersionDic

    f = open('./Passive_Update.txt', 'r')
    for line in f.readlines():
        if '=' in line:
            tempList = line.split('=')
            passiveUpdateVersionDic[tempList[0]] = tempList[1].strip()


def makePreviousPatchList():
    global previousPatchList
    global startPeriod

    f = open('./' + startPeriod.strftime('%Y_%m_%d') + '_Previous_Patch_List.txt', 'r')
    previousPatchList = f.readlines()
    f.close()


def makeSeveritySet():
    global importantSet
    global criticalSet
    global endPeriod

    xlsPath = './Security Updates ' + endPeriod.strftime('%Y-%m-%d') + '.csv'

    securityUpdates = pd.read_csv(xlsPath, encoding = 'ANSI')
    securityUpdates = securityUpdates.rename(columns={"Max Severity":"Severity"})

    for i in range(len(securityUpdates.Severity)):
        if not pd.isna(securityUpdates.Article[i]):
            if securityUpdates.Article[i].isdigit():
                if securityUpdates.Severity[i] is not None:
                    if('Important' == securityUpdates.Severity[i]):
                        importantSet.add(str(securityUpdates.Article[i]))
                    elif('Critical' == securityUpdates.Severity[i]):
                        criticalSet.add(str(securityUpdates.Article[i]))
            
    importantSet = importantSet.difference(criticalSet)


def setSeverity(excel, kbid):
    severityStr = ''
    if kbid in criticalSet:
        severityStr = '1'
    elif kbid in importantSet:
        severityStr = '0'
    return excel.replace('#s#', severityStr)


def getDownloadInfo(guid):
    url = 'https://catalog.update.microsoft.com/DownloadDialog.aspx'
    postData = {'updateIDs': '[{"size":0,"languages":"","uidInfo":"'+guid+'","updateID":"'+guid+'"}]'}
    regexForDownloadLink = "].url = '(https://.+)'"
    regexForFileName = "].fileName = '(.+)'"
    result = {}
    result['downloadLinkTuple'] = ()
    result['fileNameTuple'] = ()
    try:
        response = requests.request("POST", url, data=postData)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, "html.parser")
            headStr = str(soup.head)
            downloadLinkTuple = re.findall(regexForDownloadLink, headStr)
            result['downloadLinkTuple'] = downloadLinkTuple
            fileNameTuple = re.findall(regexForFileName, headStr)
            result['fileNameTuple'] = fileNameTuple
        else:
            None
    except requests.exceptions.Timeout as e:
        print("Timeout Error : ", e)
    except requests.exceptions.ConnectionError as e:
        print("Error Connecting : ", e)
    except requests.exceptions.HTTPError as e:
        print("Http Error : ", e)
    except requests.exceptions.RequestException as e:
        print("AnyException : ", e)

    return result


def getTargetProducts(guid):
    url = 'https://catalog.update.microsoft.com/ScopedViewInline.aspx?updateid=' + guid
    result = ''
    try:
        response = requests.request("POST", url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, "html.parser")
            result = soup.find(attrs={'id':'productsDiv'}).contents[2]
        else:
            None
    except requests.exceptions.Timeout as e:
        print("Timeout Error : ", e)
    except requests.exceptions.ConnectionError as e:
        print("Error Connecting : ", e)
    except requests.exceptions.HTTPError as e:
        print("Http Error : ", e)
    except requests.exceptions.RequestException as e:
        print("AnyException : ", e)

    return result


def makePatchRowData(guid, kbid, result, excel, descriptionInEnglish, flagForDotNet):
    # KBID, GUID, dateForm1 적용
    excel = excel.replace('#ki#', kbid).replace('#gi#', guid).replace('#df1#', endPeriod.strftime('%Y-%m-%d')).replace('#df2#', endPeriod.strftime('%Y.%m'))
    descriptionInEnglish = descriptionInEnglish.replace('#df2#', endPeriod.strftime('%Y.%m'))
    
    # 개별 변경 사항 적용
    if isinstance(result[0], str):
        excel = excel.replace('#1#', result[0])
        descriptionInEnglish = descriptionInEnglish.replace('#1#', result[0])
        if flagForDotNet:
            if '.' in result[0]:
                    versionStr = result[0].replace('및', 'and').replace('  ', ', ')
                    excel = excel.replace('#dv#', versionStr)
    else:
        # type(result[0]) == tuple 일 경우
        for i in range(len(result[0])):
            excel = excel.replace('#'+str(i+1)+'#', result[0][i])
            descriptionInEnglish = descriptionInEnglish.replace('#'+str(i+1)+'#', result[0][i])
            if flagForDotNet:
                if '.' in result[0][i]:
                        versionStr = result[0][i].replace('및', 'and').replace('  ', ', ')
                        excel = excel.replace('#dv#', versionStr)
    
    # 심각도(Severity) 적용
    excel = setSeverity(excel, kbid)

    # multi_lan용 데이터 추출
    tempList = excel.split('\t')
    tempList[8] = endPeriod.strftime('%Y년 %m월, ') + tempList[8]
    name = tempList[1]
    descr_kor = tempList[8]
    refer = tempList[9]
    korList = [name, descr_kor, refer, '']
    descriptionInEnglish = endPeriod.strftime('%B, %Y ') + descriptionInEnglish.replace('#dv#', versionStr) if flagForDotNet else endPeriod.strftime('%B, %Y ') + descriptionInEnglish
    enuList = [descriptionInEnglish, refer, '']

    # 문자열 숫자 -> 정수로 형변환
    for i in [0,3,4,10,11,15,17]:
        if tempList[i] != '':
            tempList[i] = int(tempList[i])
    return tempList, korList, enuList


def addPatchRow(Classification, guid, kbid, des, etcFlag:bool=False):
    global endPeriod
    regexList = pmsd.totalRegexDic[Classification]
    for regexDic in regexList:
        regexPattern = re.compile(regexDic['regex'].replace('#df0#', endPeriod.strftime('%Y-%m')))
        result = re.findall(regexPattern, des)
        length = len(result)
        if length == 0:
            continue
        elif length == 1:
            excel = regexDic['excel']
            descriptionInEnglish = regexDic['descriptionInEnglish']
            
            tempList, korList, enuList = makePatchRowData(guid, kbid, result, excel, descriptionInEnglish, False)

            # 확인용 데이터 추가(패치항목, 다운로드 파일 수)
            downloadInfo = getDownloadInfo(guid)
            fileNameTuple = downloadInfo['fileNameTuple']
            downloadLength = len(fileNameTuple)
            if downloadLength > 0 or isExceptionToInclude(des):
                tempList.extend([des, len(fileNameTuple)])
                # 최종행 저장
                if Classification not in pmsd.totalRowDic:
                    pmsd.totalRowDic[Classification] = {}
                if regexDic['group'] not in pmsd.totalRowDic[Classification]:
                    pmsd.totalRowDic[Classification][regexDic['group']] = []
                tempList.extend(korList + enuList)
                pmsd.totalRowDic[Classification][regexDic['group']].append(tempList)
                return
            else:
                removeList.append([guid, kbid, des])
                return
        else:
            undecidedList.append([guid, kbid, des])
            return
    if etcFlag:
        addPatchRowForMultiFile(Classification, guid, kbid, des, etcFlag)
    else:
        undecidedList.append([guid, kbid, des])


def addPatchRowForMultiFile(Classification, guid, kbid, des, etcFlag:bool=False):
    regexList = pmsd.totalRegexDicForMultiFile[Classification]
    for regexDic in regexList:
        regexPattern = re.compile(regexDic['regex'])
        result = re.findall(regexPattern, des)
        length = len(result)
        if length == 0:
            continue
        elif length == 1:
            excel = regexDic['excel']
            descriptionInEnglish = regexDic['descriptionInEnglish']
            
            tempList, korList, enuList = makePatchRowData(guid, kbid, result, excel, descriptionInEnglish, False if etcFlag else True)

            downloadInfo = getDownloadInfo(guid)
            fileNameTuple = downloadInfo['fileNameTuple']
            downloadLength = len(fileNameTuple)
            if downloadLength > 0 or isExceptionToInclude(des):
                for i in range(len(fileNameTuple)):
                    copyList = tempList.copy()
                    # 확인용 데이터 추가(패치항목, 다운로드 파일 수)
                    copyList.extend([des, 1])
                    for fileNameRegexDic in regexDic['fileName']:
                        fileNameRegexPattern = re.compile(fileNameRegexDic['regex'])
                        resultForFileName = re.findall(fileNameRegexPattern, fileNameTuple[i])
                        length = len(resultForFileName)
                        if length == 0:
                            continue
                        elif length == 1:
                            excelForFileName = fileNameRegexDic['excel']
                            if isinstance(result[0], str):
                                excelForFileName = excelForFileName.replace('#1#', result[0])
                            else:
                                for j in range(len(result[0])):
                                    excelForFileName = excelForFileName.replace('#'+str(j+1)+'#', result[0][j])
                            if isinstance(resultForFileName[0], str):
                                excelForFileName = excelForFileName.replace('#f1#', resultForFileName[0])
                            else:
                                for j in range(len(resultForFileName[0])):
                                    excelForFileName = excelForFileName.replace('#f'+str(j+1)+'#', resultForFileName[0][j])
                            copyList[16] = excelForFileName
                    #최종행 저장
                    if Classification not in pmsd.totalRowDic:
                        pmsd.totalRowDic[Classification] = {}
                    if regexDic['group'] not in pmsd.totalRowDic[Classification]:
                        pmsd.totalRowDic[Classification][regexDic['group']] = []
                    copyList.extend(korList + enuList)
                    pmsd.totalRowDic[Classification][regexDic['group']].append(copyList)
            else:
                removeList.append([guid, kbid, des])
                # tempList.extend([des, 0])
                # # 최종행 저장
                # if Classification not in pmsd.totalRowDic:
                #     pmsd.totalRowDic[Classification] = {}
                # if regexDic['group'] not in pmsd.totalRowDic[Classification]:
                #     pmsd.totalRowDic[Classification][regexDic['group']] = []
                # tempList.extend(korList + enuList)
                # pmsd.totalRowDic[Classification][regexDic['group']].append(tempList)
            return
        else:
            undecidedList.append([guid, kbid, des])
            return
    if etcFlag:
        addPatchRowByFileName(Classification, guid, kbid, des)
    else:
        undecidedList.append([guid, kbid, des])


def addPatchRowByFileName(Classification, guid, kbid, des):
    regexList = pmsd.totalRegexDicByFileName[Classification]
    flag = False
    for regexDic in regexList:
        regexPattern = re.compile(regexDic['regex'])
        result = re.findall(regexPattern, des)
        length = len(result)
        if length == 0:
            continue
        elif length == 1:
            downloadInfo = getDownloadInfo(guid)
            fileNameTuple = downloadInfo['fileNameTuple']
            downloadLength = len(fileNameTuple)
            if downloadLength > 0 or isExceptionToInclude(des):
                for fileName in fileNameTuple:
                    for one in regexDic['fileName']:
                        fileNamePattern = re.compile(one['regex'])
                        res = re.findall(fileNamePattern, fileName)
                        resLength = len(res)
                        if resLength == 0:
                            continue
                        elif resLength == 1:
                            excel = one['excel']
                            descriptionInEnglish = one['descriptionInEnglish']

                            tempList, korList, enuList = makePatchRowData(guid, kbid, result, excel, descriptionInEnglish, False)

                            # 확인용 데이터 추가(패치항목, 다운로드 파일 수)
                            tempList.extend([des, 1])

                            # 최종행 저장
                            if Classification not in pmsd.totalRowDic:
                                pmsd.totalRowDic[Classification] = {}
                            if regexDic['group'] not in pmsd.totalRowDic[Classification]:
                                pmsd.totalRowDic[Classification][regexDic['group']] = []
                            tempList.extend(korList + enuList)
                            pmsd.totalRowDic[Classification][regexDic['group']].append(tempList)
                            flag = True
                            break
            else:
                removeList.append([guid, kbid, des])
                return
        else:
            undecidedList.append([guid, kbid, des])
            return
    if not flag:
        undecidedList.append([guid, kbid, des])


def addPatchRowForMalwareRemoveTool(guid, kbid, des):
    regexList = pmsd.malwareRemoveToolRegexList
    for regexDic in regexList:
        regexPattern = re.compile(regexDic['regex'])
        result = re.findall(regexPattern, des)
        length = len(result)
        if length == 0:
            continue
        elif length == 1:
            targetProductsStr = getTargetProducts(guid)
            for one in regexDic['targetProducts']:
                if one['targetProduct'] in targetProductsStr:
                    excel = one['excel']
                    descriptionInEnglish = one['descriptionInEnglish']

                    tempList, korList, enuList = makePatchRowData(guid, kbid, result, excel, descriptionInEnglish, False)

                    # 확인용 데이터 추가(패치항목, 다운로드 파일 수)
                    tempList.extend([des, 1])

                    # 최종행 저장
                    tempList.extend(korList + enuList)
                    pmsd.totalRowDic['malware-remove-tool'][1].append(tempList)
                    return
    undecidedList.append([guid, kbid, des])


def sortPatchRowForMalwareRemoveTool():
    pmsd.totalRowDic['malware-remove-tool'][1].sort(key=lambda x:x[0])


def createMSPatchRowsByType(guid, kbid, des):
    if '.Net' in des or '.NET' in des:
        addPatchRowForMultiFile('dotnet', guid, kbid, des)
    elif 'Azure' in des:
        if 'Azure File Sync Agent' in des:
            addPatchRowByFileName('azure-file-sync-agent', guid, kbid, des)
        elif 'Azure Backup Server' in des:
            addPatchRowByFileName('azure-backup-server', guid, kbid, des)
        else:
            addPatchRow('azure', guid, kbid, des)
    elif 'Internet' in des:
        addPatchRow('internet', guid, kbid, des)
    elif 'Windows' in des:
        if '악성' in des:
            addPatchRowForMalwareRemoveTool(guid, kbid, des)
        elif any(one in des for one in ['누적', 'Cumulative']):
            addPatchRow('windows-cumulative', guid, kbid, des)
        elif any(one in des for one in ['보안', 'Security']):
            addPatchRow('windows-security', guid, kbid, des)
        else:
            addPatchRow('windows-etc', guid, kbid, des)
    elif 'Exchange' in des:
        addPatchRow('exchange', guid, kbid, des)
    elif 'PowerShell' in des:
        addPatchRow('powershell', guid, kbid, des)
    elif any(one in des for one in pmsd.officeList):
        addPatchRow('office', guid, kbid, des)
    elif 'Skype' in des:
        addPatchRow('skype', guid, kbid, des)
    elif 'SQL Server' in des:
        addPatchRowByFileName('sql-server', guid, kbid, des)
    elif 'Microsoft System Center' in des:
        addPatchRowByFileName('microsoft-system-center', guid, kbid, des)
    else:
        addPatchRow('etc', guid, kbid, des, True)


def readPatchListFromExcel():
    global endPeriod
    xlsPath = './' + endPeriod.strftime('%Y_%m_%d') + '_Result.csv'
    patchList = pd.read_csv(xlsPath, encoding = 'ANSI', names=['day', 'GUID', 'c', 'd', 'KBID', 'Des'])
    if patchList.shape[0] < 1: # == len(patchTargetList)
        print('패치 목록을 불러오는데 실패했습니다.')
        sys.exit()
    else:
        return patchList


def createMSPatchRows(patchList):
    global startPeriod
    global endPeriod
    # for i in reversed(range(patchList.shape[0])):
    for i in tqdm(range(patchList.shape[0])):
        try:
            row_datetime = dt.datetime.strptime(patchList.day[i], '%Y-%m-%dT%H:%M:%SZ').date()
            # if row_datetime >= endPeriod:
            #     continue
            # elif row_datetime < startPeriod:
            #     break
            # else:
            if row_datetime >= startPeriod and row_datetime < endPeriod:
                if validatePatchInfo(patchList.GUID[i], patchList.KBID[i], patchList.Des[i]):
                    createMSPatchRowsByType(patchList.GUID[i], str(int(patchList.KBID[i])), patchList.Des[i])
                    newPatchList.append(patchList.GUID[i]+'\t'+str(int(patchList.KBID[i]))+'\n')
        except ValueError as e: # 날짜영역에 문자열이 들어있는 경우
            # print('ValueError : ' ,e)
            None
        except TypeError as e:  # 날짜영역이 비어있는 경우
            # print('TypeError : ', e)
            None
    return None


def commonPassivePatchReplace(dic, version, replacementID, numID):
    excel = dic['excel'].replace('#df1#', endPeriod.strftime('%Y-%m-%d')).replace('#v#', version).replace('#ri#', str(replacementID))
    descriptionInEnglish = endPeriod.strftime('%B, %Y ') + dic['descriptionInEnglish'].replace('#v#', version)
    fileVersionHistory = dic['fileVersionHistory'].replace('#v#', version)
    
    excelList = excel.split('\t')
    excelList[0] = numID if numID != 0 else ''
    excelList[8] = endPeriod.strftime('%Y년 %m월, ') + excelList[8]
    name = excelList[1]
    descr_kor = excelList[8]
    refer = excelList[9]
    korList = [name, descr_kor, refer, '']
    enuList = [descriptionInEnglish, refer, '']
    for i in [0,3,4,10,11,15,17]:
        if excelList[i] != '':
            excelList[i] = int(excelList[i])
    fileVersionHistoryList = fileVersionHistory.split('\t')
    fileVersionHistoryList[0] = numID if numID != 0 else ''
    for i in [0,3,5]:
        if fileVersionHistoryList[i] != '':
            fileVersionHistoryList[i] = int(fileVersionHistoryList[i])
    return excelList, korList, enuList, fileVersionHistoryList

def makePassiveIDDic(key, value):
    if key not in passiveIDDic:
        passiveIDDic[key] = []
    passiveIDDic[key].append(value)
        

def createPassivePatchRows():
    groupDic = {'chrome':1, 'edge':2, 'adobe':3, 'hoffice2022':4, 'hoffice2020':4, 'hoffice2018':4, 'hofficeneo':4, 'hwpneo':4, 'java':5}
    global lastID
    lastID = previousIDDic['last']
    haslastID = lastID != 0
    hofficeFlag = False
    for key, value in passiveUpdateVersionDic.items():
        if value != '':
            if key == 'adobe':
                lastID = (lastID//10 + 1) * 10
                tempList = pmsd.passiveUpdateDic[key]
                for i in range(len(tempList)):
                    lastID = lastID+i if haslastID else 0
                    excelList, korList, enuList, fileVersionHistoryList = commonPassivePatchReplace(tempList[i], value, previousIDDic[key][i], lastID)
                    excelList[16] = excelList[16].replace('#vwod#', value.replace('.', ''))
                    makePassiveIDDic(key, lastID)
                    if groupDic[key] not in pmsd.totalRowDic['passive']:
                        pmsd.totalRowDic['passive'][groupDic[key]] = []
                        totalFileVersionHistoryDic[groupDic[key]] = []
                    excelList.extend(['',''] + korList + enuList)
                    pmsd.totalRowDic['passive'][groupDic[key]].append(excelList)
                    totalFileVersionHistoryDic[groupDic[key]].append(fileVersionHistoryList)
            elif key == 'java':
                lastID = (lastID//10 + 1) * 10
                tempList = pmsd.passiveUpdateDic[key]
                versionList = value.split('/')
                for i in range(len(tempList)):
                    lastID = lastID+i if haslastID else 0
                    excelList, korList, enuList, fileVersionHistoryList = commonPassivePatchReplace(tempList[i], versionList[1], previousIDDic[key][i], lastID)
                    excelList[16] = excelList[16].replace('#jv#', versionList[0])
                    makePassiveIDDic(key, lastID)
                    if groupDic[key] not in pmsd.totalRowDic['passive']:
                        pmsd.totalRowDic['passive'][groupDic[key]] = []
                        totalFileVersionHistoryDic[groupDic[key]] = []
                    excelList.extend(['',''] + korList + enuList)
                    pmsd.totalRowDic['passive'][groupDic[key]].append(excelList)
                    totalFileVersionHistoryDic[groupDic[key]].append(fileVersionHistoryList)
            else:
                if not hofficeFlag:
                    lastID = (lastID//10 + 1) * 10
                    if key in ['hoffice2022', 'hoffice2020', 'hoffice2018', 'hofficeneo', 'hwpneo']:
                        hofficeFlag = True
                else:
                    lastID = lastID + 1
                tempList = pmsd.passiveUpdateDic[key]
                for i in range(len(tempList)):
                    lastID = lastID+i if haslastID else 0
                    excelList, korList, enuList, fileVersionHistoryList = commonPassivePatchReplace(tempList[i], value, previousIDDic[key][i], lastID)
                    makePassiveIDDic(key, lastID)
                    if groupDic[key] not in pmsd.totalRowDic['passive']:
                        pmsd.totalRowDic['passive'][groupDic[key]] = []
                        totalFileVersionHistoryDic[groupDic[key]] = []
                    excelList.extend(['',''] + korList + enuList)
                    pmsd.totalRowDic['passive'][groupDic[key]].append(excelList)
                    totalFileVersionHistoryDic[groupDic[key]].append(fileVersionHistoryList)


def writePatchListToExcel():
    global endPeriod, lastID
    normalRowCount = 0
    fileVerHistoryRowCount = 0
    wb = Workbook()
    normal_ws = wb.active
    normal_ws.title = 'normal'
    for key, value in pmsd.totalRowDic.items():
        value = dict(sorted(value.items()))
        for list in value.values():
            if key != 'passive':
                list.sort(key=lambda x:x[8])
                if key != 'malware-remove-tool':
                    lastID = (lastID//10 + 1) * 10
            for one in list:
                if one[0] == '':
                    one[0] = lastID
                    lastID += 1
                normal_ws.append(one)
            if key != 'passive' and key != 'malware-remove-tool':
                lastID -= 1
            normal_ws.append([])
            normal_ws.append([])
            normalRowCount += len(list) + 2

    undecided_ws = wb.create_sheet()
    undecided_ws.title = 'undecided'
    for one in undecidedList:
        undecided_ws.append(one)

    duplication_ws = wb.create_sheet()
    duplication_ws.title = 'duplication'
    for one in duplicationPatchList:
        duplication_ws.append(one)

    exclusion_ws = wb.create_sheet()
    exclusion_ws.title = 'exclusion'
    for one in exclusionPatchList:
        exclusion_ws.append(one)

    remove_ws = wb.create_sheet()
    remove_ws.title = 'remove'
    for one in removeList:
        remove_ws.append(one)

    file_ver_history_ws = wb.create_sheet()
    file_ver_history_ws.title = 'file_ver_history'
    for list in totalFileVersionHistoryDic.values():
        for one in list:
            file_ver_history_ws.append(one)
        file_ver_history_ws.append([])
        file_ver_history_ws.append([])
        fileVerHistoryRowCount += len(list) + 2

    xlsxPath = './' + endPeriod.strftime('%Y_%m_%d') + '_Auto_Patch.xlsx'
    # DEFAULT_FONT.name = '굴림체'
    DEFAULT_FONT.size = '10'
    for row in normal_ws['1:' + str(normalRowCount)]:
        for cell in row:
            cell.font = Font(name='굴림체')
    for row in file_ver_history_ws['1:' + str(fileVerHistoryRowCount)]:
        for cell in row:
            cell.font = Font(name='굴림체')
    wb.save(xlsxPath)

def removeListFromNewPatchList():
    for one in removeList:
        newPatchList.remove(one[0]+'\t'+str(int(one[1]))+'\n')

def writePreviousPatchListTxt():
    removeListFromNewPatchList()
    f = open('./' + endPeriod.strftime('%Y_%m_%d') + '_Previous_Patch_List.txt', 'w')
    totalPatchList = previousPatchList + newPatchList
    for one in totalPatchList:
        f.write(one)
    f.close()

def writePreviousPassiveUpateTxt():
    global lastID
    f = open('./' + endPeriod.strftime('%Y_%m_%d') + '_Previous_Passive_Update.txt', 'w')
    for key, value in passiveIDDic.items():
        previousIDDic[key] = value
    previousIDDic['last'] = lastID
    f.write('previousID=' + str(previousIDDic))
    f.close()

def main():
    global startPeriod
    global endPeriod
    numberOfArgs = len(sys.argv)
    if numberOfArgs == 1:
        getPatchPeriod()
    elif numberOfArgs == 3:
        startPeriod, endPeriod = dt.datetime.strptime(sys.argv[1], '%Y%m%d').date(), dt.datetime.strptime(sys.argv[2], '%Y%m%d').date()
    else:
        print('파라미터 개수를 확인해주세요.')
        sys.exit()

    makeSeveritySet()
    makePreviousPatchList()
    makePreviousIDDic()
    makePassiveUpdateVersionDic()

    createPassivePatchRows()
    patchList = readPatchListFromExcel()
    createMSPatchRows(patchList)
    sortPatchRowForMalwareRemoveTool()

    writePatchListToExcel()
    writePreviousPatchListTxt()
    writePreviousPassiveUpateTxt()

startTime = time.time()
main()
endTime = time.time()
print(f'소요시간 : {endTime-startTime:.3f} sec')